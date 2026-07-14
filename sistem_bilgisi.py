# -*- coding: utf-8 -*-
"""
Sistem Bilgisi Toplayici
Staj Projesi - 1., 2., 3. ve 4. Asama (WMI ve WinReg Optimizasyonlu)

Bu script;
  - Bilgisayar adi, isletim sistemi, CPU, RAM ve disk bilgilerini,
  - Yuklu program kontrolunu (Chrome, Office, Antivirus) (winreg & wmi ile),
  - IP durumu (DHCP / Statik) sorgusunu (wmi ile),
  - Standart kontrol mantigina gore degerlendirme uyarilarini
ekrana yazdirir ve pc_rapor_{bilgisayar_adi}.xlsx olarak disa aktarir.
"""

import datetime
import os
import platform
import socket
import sys
import io
import psutil
import openpyxl
import wmi
import winreg
import pythoncom

def with_com(func):
    """WMI fonksiyonlarini ayri bir is parcaciginda cagirirken COM nesnelerini guvenle baslatir."""
    def wrapper(*args, **kwargs):
        pythoncom.CoInitialize()
        try:
            return func(*args, **kwargs)
        finally:
            pythoncom.CoUninitialize()
    return wrapper

# Not: stdout UTF-8 yonlendirmesi yalnizca dogrudan calistirildigi zaman
# etkinlestirilir (asagida __main__ blogu icinde). Boylece arayuz.py
# tarafindan import edildiginde tkinter boru hattini bozmaz.

def bytes_to_gb(bayt: int) -> float:
    """Bayt cinsinden gelen değeri GB'a çevirir."""
    return round(bayt / (1024 ** 3), 2)

def bilgisayar_adi() -> str:
    """Bilgisayarın ağ adını döndürür."""
    return socket.gethostname()

def isletim_sistemi_bilgisi() -> dict:
    """İşletim sistemi bilgilerini döndürür."""
    return {
        "Sistem"   : platform.system(),
        "Sürüm"    : platform.version(),
        "Mimari"   : platform.machine(),
        "İşlemci"  : platform.processor(),
    }

def cpu_bilgisi() -> dict:
    """CPU çekirdek sayısı ve anlık kullanım yüzdesini döndürür."""
    return {
        "Fiziksel Çekirdek Sayısı" : psutil.cpu_count(logical=False),
        "Mantıksal Çekirdek Sayısı": psutil.cpu_count(logical=True),
        "Mevcut Kullanım (%)"      : psutil.cpu_percent(interval=1),
        "Maksimum Frekans (MHz)"   : psutil.cpu_freq().max if psutil.cpu_freq() else "Bilinmiyor",
        "Mevcut Frekans (MHz)"     : psutil.cpu_freq().current if psutil.cpu_freq() else "Bilinmiyor",
    }

def ram_bilgisi() -> dict:
    """Toplam, kullanılan ve boş RAM miktarını döndürür."""
    ram = psutil.virtual_memory()
    return {
        "Toplam RAM (GB)"   : bytes_to_gb(ram.total),
        "Kullanılan (GB)"   : bytes_to_gb(ram.used),
        "Boş (GB)"          : bytes_to_gb(ram.available),
        "Kullanım Oranı (%)": ram.percent,
    }

def disk_bilgisi() -> list[dict]:
    """Tüm disk bölümlerinin doluluk bilgisini döndürür."""
    disk_listesi = []
    for bolum in psutil.disk_partitions(all=False):
        try:
            kullanim = psutil.disk_usage(bolum.mountpoint)
            disk_listesi.append({
                "Bağlama Noktası"    : bolum.mountpoint,
                "Dosya Sistemi"      : bolum.fstype,
                "Toplam (GB)"        : bytes_to_gb(kullanim.total),
                "Kullanılan (GB)"    : bytes_to_gb(kullanim.used),
                "Boş (GB)"           : bytes_to_gb(kullanim.free),
                "Doluluk Oranı (%)"  : kullanim.percent,
            })
        except PermissionError:
            continue
    return disk_listesi

@with_com
def ip_durumu() -> str:
    """
    Donanımsal olarak var olan TÜM fiziksel ağ kartlarını tarar.
    Herhangi bir fiziksel kartta DHCP kapalıysa (statik IP), kart
    o an bağlı olmasa bile 'Statik IP (Manuel)' döner.

    Filtreleme:
      - Win32_NetworkAdapter(PhysicalAdapter=True) ile donanımsal kartlar alınır.
      - Ek güvenlik için Description/Caption üzerinde sanal kelime filtresi uygulanır.
      - DHCPEnabled=None (yapılandırma kaydı olmayan pasif kartlar) DHCP kabul edilir.
    """
    SANAL_KELIMELER = [
        "vmware", "virtualbox", "loopback", "bluetooth",
        "vpn", "vethernet", "hyper-v", "virtual", "tunnel",
        "teredo", "isatap", "6to4", "microsoft wi-fi direct",
    ]

    def _sanal_mi(metin: str) -> bool:
        metin = metin.lower()
        return any(k in metin for k in SANAL_KELIMELER)

    try:
        c = wmi.WMI()

        # PhysicalAdapter=True → donanımsal kartlar (sanal/tünel hariç)
        fiziksel_nicler = c.Win32_NetworkAdapter(PhysicalAdapter=True)

        statik_bulundu    = False
        fiziksel_kart_var = False

        for nic in fiziksel_nicler:
            # İkinci güvenlik katmanı: açıklama/ad üzerinden sanal filtresi
            desc    = str(getattr(nic, "Description",    "") or "")
            caption = str(getattr(nic, "Caption",        "") or "")
            if _sanal_mi(desc + " " + caption):
                continue

            # Bu karta ait WMI yapılandırma kaydını al
            configs = c.Win32_NetworkAdapterConfiguration(Index=nic.Index)
            for cfg in configs:
                fiziksel_kart_var = True
                dhcp = getattr(cfg, "DHCPEnabled", None)
                # DHCPEnabled=False  → Statik (DHCP kapalı)
                # DHCPEnabled=True   → Otomatik
                # DHCPEnabled=None   → Kart pasif/bağlantısız, yapılandırma yok → DHCP say
                if dhcp is False:
                    statik_bulundu = True
                    break

            if statik_bulundu:
                break   # Bir tane bile statik kart yeterlii

        if not fiziksel_kart_var:
            return "Bilinmiyor"

        return "Statik IP (Manuel)" if statik_bulundu else "Otomatik IP (DHCP)"

    except Exception as e:
        return f"Hata Olustu: {str(e)}"


@with_com
def ag_uyeligi() -> str:
    """
    WMI Win32_ComputerSystem ile bilgisayarin Domain veya Workgroup
    uyeligini sorgular.
    """
    try:
        c = wmi.WMI()
        for sistem in c.Win32_ComputerSystem():
            if getattr(sistem, "PartOfDomain", False):
                return f"Etki Alani: {getattr(sistem, 'Domain', 'Bilinmiyor')}"
            else:
                return f"Calisma Grubu: {getattr(sistem, 'Workgroup', 'Bilinmiyor')}"
        return "Bilinmiyor"
    except Exception as e:
        return f"Hata Olustu: {str(e)}"


# ──────────────────────────────────────────────────────────────────────────────
# AŞAMA 2 – YÜKLÜ PROGRAM KONTROLÜ (WinReg & WMI)
# ──────────────────────────────────────────────────────────────────────────────

def _kayit_defterinde_ara(aranacak_kelimeler: list[str]) -> bool:
    """
    Kayıt defterindeki Uninstall dizinlerinde verilen kelimeleri arar.
    """
    yollar = [
        r"SOFTWARE\Microsoft\Windows\CurrentVersion\Uninstall",
        r"SOFTWARE\WOW6432Node\Microsoft\Windows\CurrentVersion\Uninstall"
    ]
    hives = [winreg.HKEY_LOCAL_MACHINE, winreg.HKEY_CURRENT_USER]
    
    for hive in hives:
        for yol in yollar:
            try:
                key = winreg.OpenKey(hive, yol)
                for i in range(0, winreg.QueryInfoKey(key)[0]):
                    try:
                        subkey_name = winreg.EnumKey(key, i)
                        subkey = winreg.OpenKey(key, subkey_name)
                        name = winreg.QueryValueEx(subkey, "DisplayName")[0]
                        
                        for kelime in aranacak_kelimeler:
                            if kelime.lower() in str(name).lower():
                                winreg.CloseKey(subkey)
                                winreg.CloseKey(key)
                                return True
                        winreg.CloseKey(subkey)
                    except EnvironmentError:
                        continue
                winreg.CloseKey(key)
            except Exception:
                continue
    return False

def chrome_kurulu_mu() -> str:
    """Google Chrome'un kurulu olup olmadigini kontrol eder."""
    return "VAR" if _kayit_defterinde_ara(["Google Chrome"]) else "YOK"

def office_kurulu_mu() -> str:
    """Microsoft Office'in kurulu olup olmadigini kontrol eder."""
    return "VAR" if _kayit_defterinde_ara(["Microsoft Office", "Microsoft 365", "Office 365"]) else "YOK"

def teamviewer_kurulu_mu() -> str:
    """
    TeamViewer'in kurulu olup olmadigini kayit defteri ve
    varsayilan kurulum yolu uzerinden kontrol eder.
    """
    # Oncelikle kayit defterinde ara
    if _kayit_defterinde_ara(["TeamViewer"]):
        return "VAR"
    # Ikincil kontrol: varsayilan kurulum dizinleri
    kurulum_yollari = [
        r"C:\Program Files\TeamViewer\TeamViewer.exe",
        r"C:\Program Files (x86)\TeamViewer\TeamViewer.exe",
    ]
    for yol in kurulum_yollari:
        if os.path.exists(yol):
            return "VAR"
    return "YOK"

@with_com
def bitdefender_kurulu_mu() -> str:
    """
    Kurumsal standart antivirüs olan 'Bitdefender Endpoint Security Tools'
    programinin kurulu olup olmadigini kontrol eder.
    Oncelik sirasi:
      1. Kayit defteri (tum Hive'lar ve Uninstall anahtarlari)
      2. Bilinen kurulum dizinleri
      3. WMI SecurityCenter2 (aktif urun adi eslesimi)
    """
    ARAMA_KELIMELERI = ["Bitdefender Endpoint", "Bitdefender Endpoint Security"]
    KURULUM_YOLLARI = [
        r"C:\Program Files\Bitdefender\Endpoint Security\EPSecurityService.exe",
        r"C:\Program Files\Bitdefender\Bitdefender Endpoint Security\EPSecurityService.exe",
        r"C:\Program Files (x86)\Bitdefender\Endpoint Security\EPSecurityService.exe",
    ]

    # 1. Kayit defteri kontrolu
    if _kayit_defterinde_ara(ARAMA_KELIMELERI):
        return "VAR"

    # 2. Kurulum dizini kontrolu
    for yol in KURULUM_YOLLARI:
        if os.path.exists(yol):
            return "VAR"

    # 3. WMI SecurityCenter2 kontrolu
    try:
        c_sc = wmi.WMI(namespace=r"root\SecurityCenter2")
        for av in c_sc.AntiVirusProduct():
            ad = str(getattr(av, "displayName", "") or "").strip().lower()
            if "bitdefender" in ad:
                return "VAR"
    except Exception:
        pass

    return "YOK"


def acrobat_reader_kurulu_mu() -> str:
    """
    Adobe Acrobat Reader'in kurulu olup olmadigini kontrol eder.
    Oncelik: Kayit defteri → bilinen kurulum yollari.
    """
    ARAMA_KELIMELERI = ["Adobe Acrobat", "Acrobat Reader"]
    KURULUM_YOLLARI = [
        r"C:\Program Files\Adobe\Acrobat DC\Acrobat\Acrobat.exe",
        r"C:\Program Files (x86)\Adobe\Acrobat Reader DC\Reader\AcroRd32.exe",
        r"C:\Program Files\Adobe\Acrobat Reader DC\Reader\AcroRd32.exe",
        r"C:\Program Files (x86)\Adobe\Reader 11.0\Reader\AcroRd32.exe",
    ]
    if _kayit_defterinde_ara(ARAMA_KELIMELERI):
        return "VAR"
    for yol in KURULUM_YOLLARI:
        if os.path.exists(yol):
            return "VAR"
    return "YOK"


def sikistirma_araci_kurulu_mu() -> str:
    """
    WinRAR, 7-Zip veya WinZip'ten herhangi birinin kurulu olup olmadigini kontrol eder.
    Bulunan ilk aracin adini dondurur; hicbiri yoksa 'YOK' dondurur.
    """
    ARACLAR = [
        ("WinRAR",  ["WinRAR"],  [
            r"C:\Program Files\WinRAR\WinRAR.exe",
            r"C:\Program Files (x86)\WinRAR\WinRAR.exe",
        ]),
        ("7-Zip",   ["7-Zip"],   [
            r"C:\Program Files\7-Zip\7z.exe",
            r"C:\Program Files (x86)\7-Zip\7z.exe",
        ]),
        ("WinZip",  ["WinZip"],  [
            r"C:\Program Files\WinZip\WinZip64.exe",
            r"C:\Program Files (x86)\WinZip\WinZip32.exe",
        ]),
    ]
    for ad, kelimeler, yollar in ARACLAR:
        if _kayit_defterinde_ara(kelimeler):
            return f"VAR ({ad})"
        for yol in yollar:
            if os.path.exists(yol):
                return f"VAR ({ad})"
    return "YOK"


@with_com
def windows_lisans_durumu() -> str:
    """
    WMI SoftwareLicensingProduct uzerinden Windows aktivasyon durumunu sorgular.
    LicenseStatus == 1 → Lisansli (Aktif).
    """
    WINDOWS_APP_ID = "55c92734-d682-4d71-983e-d6ec3f16059f"
    try:
        c = wmi.WMI()
        urunler = c.SoftwareLicensingProduct(
            ApplicationId=WINDOWS_APP_ID
        )
        for u in urunler:
            durum = getattr(u, "LicenseStatus", None)
            if durum == 1:
                return "Lisansli (Aktif)"
        return "Lisanssiz / Aktif Degil"
    except Exception as e:
        return f"Sorgulanamadi ({e})"


@with_com
def office_lisans_durumu() -> str:
    """
    WMI SoftwareLicensingProduct uzerinden Microsoft Office aktivasyon durumunu sorgular.
    Birden fazla Office urunu varsa, aktif olan herhangi biri yeterlidir.
    """
    OFFICE_APP_IDS = [
        "0ff1ce15-a989-479d-af46-f275c6370663",  # Office 2016/2019/365 ortak
        "59a52881-a989-479d-af46-f275c6370663",  # Office 2013
    ]
    try:
        c = wmi.WMI()
        for app_id in OFFICE_APP_IDS:
            try:
                urunler = c.SoftwareLicensingProduct(ApplicationId=app_id)
                for u in urunler:
                    ad    = str(getattr(u, "Name",          "") or "")
                    durum = getattr(u, "LicenseStatus", None)
                    if durum == 1 and "office" in ad.lower():
                        return "Lisansli"
            except Exception:
                continue
        return "Lisanssiz / Aktif Degil"
    except Exception as e:
        return f"Sorgulanamadi ({e})"


def yuklu_program_kontrol() -> dict:
    # Office kurulum + lisans bilgisini tek geciste topla
    office_kurulum = office_kurulu_mu()
    office_lisans  = office_lisans_durumu() if "VAR" in office_kurulum else "Kurulu Degil"
    return {
        "Google Chrome"     : chrome_kurulu_mu(),
        "Microsoft Office"  : f"{office_kurulum}|{office_lisans}",
        "TeamViewer"        : teamviewer_kurulu_mu(),
        "Antivirus"         : bitdefender_kurulu_mu(),
        "Acrobat Reader"    : acrobat_reader_kurulu_mu(),
        "Sikistirma Araci"  : sikistirma_araci_kurulu_mu(),
        "Windows Lisans"    : windows_lisans_durumu(),
    }


# ──────────────────────────────────────────────────────────────────────────────
# ASAMA 3 – STANDART KONTROL MANTIGI (DEGERLENDIRME)
# ──────────────────────────────────────────────────────────────────────────────

RAM_ESIK_GB       = 8.0
DISK_ESIK_YUZDE  = 85.0

def standart_kontrol(
    ram_gb: float,
    office_durum: str,
    disk_yuzde: float,
    teamviewer_durum: str = "",
    ip: str = "",
    uyelik: str = "",
    antivirus_durum: str = "",
    acrobat_durum: str = "",
    sikistirma_durum: str = "",
    windows_lisans: str = "",
) -> list[str]:
    """
    Donanim, yazilim ve ag kriterlerini degerlendirerek uyari listesi dondurur.

    Kurallar:
      - RAM < 8 GB
      - Microsoft Office kurulu degil veya lisanssiz
      - Disk dolulugu > %85
      - TeamViewer yok
      - Bitdefender Endpoint Security Tools yok
      - Adobe Acrobat Reader yok
      - Sikistirma araci yok
      - Windows lisanssiz
      - Statik / Manuel IP konfigurasyonu
      - Cihaz etki alani (Domain) disinda (Workgroup)
    """
    uyarilar: list[str] = []

    # ── Donanim ──────────────────────────────────────────────────────────────
    if ram_gb < RAM_ESIK_GB:
        uyarilar.append(f"[!] RAM yetersiz: {ram_gb} GB < {RAM_ESIK_GB} GB esik")
    if disk_yuzde > DISK_ESIK_YUZDE:
        uyarilar.append(f"[!] Disk dolulugu kritik: %{disk_yuzde} > %{DISK_ESIK_YUZDE} esik")

    # ── Yazilim ──────────────────────────────────────────────────────────────
    # Office: 'VAR|Lisanssiz' veya 'YOK|...' her ikisi de uyari olusturur
    if "YOK" in office_durum.upper():
        uyarilar.append("[!] Microsoft Office kurulu degil")
    elif "Lisanssiz" in office_durum or "Aktif Degil" in office_durum:
        uyarilar.append("[!] Microsoft Office lisanssiz / aktif degil")
    if teamviewer_durum and "YOK" in teamviewer_durum.upper():
        uyarilar.append("[-] TeamViewer kurulu degil")
    if antivirus_durum and "YOK" in antivirus_durum.upper():
        uyarilar.append("[-] Bitdefender Endpoint Security Tools kurulu degil")
    if acrobat_durum and "YOK" in acrobat_durum.upper():
        uyarilar.append("[-] Adobe Acrobat Reader kurulu degil")
    if sikistirma_durum and "YOK" in sikistirma_durum.upper():
        uyarilar.append("[-] Sikistirma araci (WinRAR/7-Zip/WinZip) kurulu degil")
    if windows_lisans and ("Lisanssiz" in windows_lisans or "Aktif Degil" in windows_lisans):
        uyarilar.append("[!] Windows lisanssiz / aktif degil")

    # ── Ag ───────────────────────────────────────────────────────────────────
    if ip and "Statik" in ip:
        uyarilar.append("[-] Statik IP yapilandirmasi mevcut")
    if uyelik and ("Calisma Grubu" in uyelik or "Workgroup" in uyelik.lower()):
        uyarilar.append("[-] Cihaz Domain'e bagli degil")

    return uyarilar


# ──────────────────────────────────────────────────────────────────────────────
# ASAMA 4 – EXCEL RAPORU
# ──────────────────────────────────────────────────────────────────────────────

def excel_raporu_kaydet(
    ram: dict, cpu: dict, isletim_sistemi: dict, diskler: list[dict],
    programlar: dict, uyarilar: list[str], ip: str = "",
    ag_uyeligi_bilgi: str = "",
    cikti_dizini: str = ".", dosya_adi: str = ""
) -> str:
    zaman_damgasi = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    bilgisayar = bilgisayar_adi()

    if not dosya_adi:
        temiz_ad = "".join(c if c.isalnum() or c in ("-", "_") else "_" for c in bilgisayar)
        dosya_adi = f"pc_rapor_{temiz_ad}.xlsx"
        
    ozet_satirlar = [
        {"Kategori": "Rapor Zamani",       "Bilgi": zaman_damgasi},
        {"Kategori": "Bilgisayar Adi",     "Bilgi": bilgisayar},
        {"Kategori": "IP Durumu",          "Bilgi": ip if ip else "Bilinmiyor"},
        {"Kategori": "Ag Uyeligi",         "Bilgi": ag_uyeligi_bilgi if ag_uyeligi_bilgi else "Bilinmiyor"},
        {"Kategori": "Isletim Sistemi",    "Bilgi": isletim_sistemi.get("Sistem", "")},
        {"Kategori": "OS Surumu",          "Bilgi": isletim_sistemi.get("Sürüm", "")},
        {"Kategori": "Mimari",             "Bilgi": isletim_sistemi.get("Mimari", "")},
        {"Kategori": "Islemci",            "Bilgi": isletim_sistemi.get("İşlemci", "")},
        {"Kategori": "Fiziksel Cekirdek",  "Bilgi": cpu.get("Fiziksel Çekirdek Sayısı", "")},
        {"Kategori": "Mantiksal Cekirdek", "Bilgi": cpu.get("Mantıksal Çekirdek Sayısı", "")},
        {"Kategori": "CPU Kullanim (%)",   "Bilgi": cpu.get("Mevcut Kullanım (%)", "")},
        {"Kategori": "Toplam RAM (GB)",    "Bilgi": ram.get("Toplam RAM (GB)", "")},
        {"Kategori": "Kullanilan RAM (GB)","Bilgi": ram.get("Kullanılan (GB)", "")},
        {"Kategori": "Bos RAM (GB)",       "Bilgi": ram.get("Boş (GB)", "")},
        {"Kategori": "RAM Kullanim (%)",   "Bilgi": ram.get("Kullanım Oranı (%)", "")},
    ]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_ozet = wb.create_sheet("Sistem Ozeti")
    ws_ozet.append(["Kategori", "Bilgi"])
    for row in ozet_satirlar:
        ws_ozet.append([row["Kategori"], row["Bilgi"]])

    ws_disk = wb.create_sheet("Disk Bilgileri")
    ws_disk.append(["Baglama Noktasi", "Dosya Sistemi", "Toplam (GB)", "Kullanilan (GB)", "Bos (GB)", "Doluluk Orani (%)"])
    if diskler:
        for d in diskler:
            ws_disk.append([
                d.get("Baglama Noktasi", ""), d.get("Dosya Sistemi", ""), 
                d.get("Toplam (GB)", ""), d.get("Kullanilan (GB)", ""), 
                d.get("Bos (GB)", ""), d.get("Doluluk Orani (%)", "")
            ])

    ws_program = wb.create_sheet("Program Durumu")
    ws_program.append(["Program", "Durum"])
    for prog, durum in programlar.items():
        ws_program.append([prog, durum])

    ws_uyari = wb.create_sheet("Degerlendirme")
    ws_uyari.append(["Uyari"])
    if uyarilar:
        for u in uyarilar:
            ws_uyari.append([u.lstrip("[!] ").strip()])
    else:
        ws_uyari.append(["Tum kriterler karsilandi – uyari yok."])

    dosya_yolu = os.path.join(cikti_dizini, dosya_adi)
    
    for sayfa in wb.worksheets:
        for sutun_hucreleri in sayfa.columns:
            max_uzunluk = max(
                len(str(hucre.value)) if hucre.value is not None else 0
                for hucre in sutun_hucreleri
            )
            sayfa.column_dimensions[sutun_hucreleri[0].column_letter].width = min(max_uzunluk + 4, 60)

    wb.save(dosya_yolu)
    return os.path.abspath(dosya_yolu)


def raporu_yazdir() -> None:
    sep2 = "-" * 48

    print("[YUKLU PROGRAM KONTROLU]")
    print(sep2)
    programlar = yuklu_program_kontrol()
    _program_etiket = {
        "Google Chrome"    : "Google Chrome",
        "Microsoft Office" : "Microsoft Office",
        "TeamViewer"       : "TeamViewer",
        "Antivirus"        : "Antivirus: Bitdefender Endpoint Security Tools",
    }
    for prog, durum in programlar.items():
        etiket = _program_etiket.get(prog, prog)
        if "VAR" in durum:
            print(f"[\u2713] {etiket}")
        else:
            print(f"[X] {etiket} (Bulunamadi)")

    print()
    print("[AG / IP DURUMU]")
    print(sep2)
    ip     = ip_durumu()
    uyelik = ag_uyeligi()
    print(f"IP Yapilandirmasi  : {ip}")
    print(f"Ag Uyeligi         : {uyelik}")

    # Excel raporu: arka planda hesaplanip kaydedilir
    ram_veri   = ram_bilgisi()
    disk_veri  = disk_bilgisi()
    disk_yuzde = next((d["Doluluk Oranı (%)"] for d in disk_veri if d["Bağlama Noktası"] == "C:\\"), 0.0)
    uyarilar   = standart_kontrol(
        ram_gb           = ram_veri["Toplam RAM (GB)"],
        office_durum     = programlar["Microsoft Office"],
        disk_yuzde       = disk_yuzde,
        teamviewer_durum = programlar.get("TeamViewer", ""),
        ip               = ip,
        uyelik           = uyelik,
        antivirus_durum  = programlar.get("Antivirus", ""),
    )
    yol = excel_raporu_kaydet(
        ram              = ram_veri,
        cpu              = cpu_bilgisi(),
        isletim_sistemi  = isletim_sistemi_bilgisi(),
        diskler          = disk_veri,
        programlar       = programlar,
        uyarilar         = uyarilar,
        ip               = ip,
        ag_uyeligi_bilgi = uyelik,
    )
    print(f"\n[Rapor kaydedildi] {yol}")



if __name__ == "__main__":
    if sys.platform == "win32":
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    raporu_yazdir()